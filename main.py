# ike.gabrielyan@yandex.ru
import datetime
import json
from math import ceil, floor
import vk_api
import openpyxl
import re
import PIL.Image as Image

from range_key_dict import RangeKeyDict
from vk_api.longpoll import VkLongPoll, VkEventType
from vk_api.keyboard import VkKeyboard, VkKeyboardColor
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import matplotlib
import requests
import strings
from bs4 import BeautifulSoup

TOKEN = strings.TOKEN

vk_session = vk_api.VkApi(token=TOKEN)
vk = vk_session.get_api()
longpoll = VkLongPoll(vk_session)
group_pattern = r'^[и]{1}[а-я]{3}[-]{1}\d{2}[-]{1}\d{2}$'


# Класс пользователя
class User:
    def __init__(self, id, mode, group):
        self.id = id
        self.mode = mode
        self.group = group
        self.search_teacher = ''
        self.search_group = group

users = []
with open('users.json', 'r') as f:
    json_users = json.load(f)
    users_dumped = dict(json_users) # список пользователей
    users_dumped = users_dumped.get("users")
    if(users_dumped is not None):
        for user in users_dumped:
            user = dict(user)
            users.append(User(user.get("id"), user.get("mode"), user.get("group")))


empty_keyboard = {'keyboard': {
    "one_time": False,
    "buttons": []
}}

# день недели
weekday_list = ["понедельник", 'вторник', 'среда', 'четверг', 'пятница', 'суббота', 'воскресенье']
weekday = {1: "понедельник", 2: 'вторник', 3: 'среда', 4: 'четверг', 5: 'пятница', 6: 'суббота', 7: 'воскресенье'}

# время суток
day_time = RangeKeyDict({(5, 11): 0, (11, 16): 1, (16, 23): 2, (23, 24): 3, (0, 5): 3})
day_time_name = {0: "утро", 1: "день", 2: "вечер", 3: "ночь"}

# характеристика ветра
wind_type = RangeKeyDict({(0, 0.3): 'штиль', (0.3, 1.6): 'тихий',
                          (1.6, 3.4): 'легкий', (3.4, 5.5): 'слабый', (5.5, 8): 'умеренный',
                          (8, 10.8): 'свежий', (10.8, 13.9): 'сильный', (13.9, 17.2): 'крепкий',
                          (17.2, 20.8): 'очень крепкий', (20.8, 24.5): 'шторм', (24.5, 28.5): 'сильный шотрм',
                          (28.5, 32.7): 'жестокий шторм', (32.7, 200): 'ураган'})
wind_direction = RangeKeyDict({(0, 22.5): 'северный',
                               (22.5, 67.5): 'северо-восточный',
                               (67.5, 112.5): 'восточный',
                               (112.5, 157.5): 'юго-восточный',
                               (157.5, 202.5): 'южный',
                               (202.5, 247.5): 'юго-западный',
                               (247.5, 292.5): 'западный',
                               (292.5, 337.5): 'северо-западный',
                               (337.5, 360.1): 'северный'})
# словарь: название региона - ссылка на стр со статистикой
reg_dict = {}
reg_list = []
page_reg = requests.get('https://coronavirusstat.ru/country/russia/')  # адрес страницы со статистикой
region_soup = BeautifulSoup(page_reg.text, "html.parser")
a = (region_soup.find_all('div', class_="p-1 col-5"))
for i in a:
    reg_list.append(' '.join(i.text.split()[1:]))
    reg_dict[' '.join(i.text.split()[1:])] = 'https://coronavirusstat.ru' + i.find('a').get('href')


# отправка текстового сообщения
def send_message(user_id, message, keyboard=None):
    post = {'user_id': user_id, 'random_id': 0, 'message': message}
    if keyboard != None:
        post['keyboard'] = keyboard.get_keyboard()
    vk_session.method('messages.send', post)


# отправка сообщения с фотографией во вложениях
def send_photo(user_id, img_req, message=None):
    upload = vk_api.VkUpload(vk_session)
    photo = upload.photo_messages(img_req)[0]
    owner_id = photo['owner_id']
    photo_id = photo['id']
    attachment = f'photo{owner_id}_{photo_id}'
    post = {'user_id': user_id, 'random_id': 0, "attachment": attachment}
    if message != None:
        post['message'] = message
    vk_session.method('messages.send', post)



def get_schedule_for_course(course_num):
    page_mirea = requests.get('https://www.mirea.ru/schedule/')
    soup_mirea = BeautifulSoup(page_mirea.text, 'html.parser')
    result = soup_mirea.find("div", class_='rasspisanie').find(string="Институт информационных технологий").find_parent(
        "div").find_parent("div")
    result = result.find_all("a", class_='uk-link-toggle')[course_num - 1].get('href')
    with open(f"file_ras{course_num}.xlsx", "wb") as f:
        resp = requests.get(result)
        f.write((resp.content))
    book = openpyxl.load_workbook(f"file_ras{course_num}.xlsx")
    sheet = book.active
    return sheet


#schedule_list = [get_schedule_for_course(1), get_schedule_for_course(2), get_schedule_for_course(3)]

book = openpyxl.load_workbook("raspisanie1.xlsx")
schedule1 = book.active
book = openpyxl.load_workbook("raspisanie2.xlsx")
schedule2 = book.active
book = openpyxl.load_workbook("raspisanie3.xlsx")
schedule3 = book.active
schedule_list = [schedule1, schedule2, schedule3]  # пока открываем загруженное старое расписание т.к. зачетная неделя

# расписание на любой день
def schedule_for_any_day(gr_num, d=datetime.datetime.today()):
    course_num = int(d.strftime('%y')) - int(gr_num[8:]) + int(d.month > 8)
    if course_num > 3 or course_num < 0:
        string_to_post = "Данной группы нет в расписании"
        return string_to_post
    sheet = schedule_list[course_num - 1]
    num_cols = sheet.max_column
    gr_num = str(gr_num).upper()
    num_rows = sheet.max_row
    group_cell = None
    wn = int(d.isoweekday())
    string_to_post = ''
    if wn == 7:
        string_to_post = "Выходной!"
        return string_to_post
    for i in range(6, num_cols + 1, 5):
        if (sheet.cell(row=2, column=i).value == gr_num):
            group_cell = [2, i]
            break
    if group_cell == None:
        string_to_post = "Нет такой группы в расписании :("
        return string_to_post
    sub_cell = [4 + (wn - 1) * 12 + (week_number(d) + 1) % 2, i]
    for i in range(6):
        val = sheet.cell(sub_cell[0] + i * 2, sub_cell[1]).value
        val_inf = [sheet.cell(sub_cell[0] + i * 2, sub_cell[1] + 1).value,
                   sheet.cell(sub_cell[0] + i * 2, sub_cell[1] + 2).value,
                   sheet.cell(sub_cell[0] + i * 2, sub_cell[1] + 3).value]
        if val != None and str(val).lower() != str(val).upper():
            string_to_post += str(i + 1) + ') '
            for t in range(len(val.split('\n'))):
                string_to_post += str(val).split('\n')[t]
                for y in range(3):
                    if val_inf[y] == None:
                        string_to_post += ", ---"
                    else:
                        try:
                            string_to_post += ", " + str(val_inf[y]).split('\n')[t]
                        except IndexError:
                            pass
                string_to_post += '\n'
    if string_to_post == "":
        string_to_post = "Выходной!"
    string_to_post = "Расписание на " + d.strftime("%d.%m") + '\n' + string_to_post
    return string_to_post


# расписание на dw день недли
def schedule_week_day(dw, gr_num):
    ndw = None
    gr_num = gr_num.upper()
    str, s = '', ['', '']
    for i in weekday:
        if weekday[i] == dw:
            ndw = i
    if ndw == None:
        str = "Ошибка в названии дня недели!"
        return str
    course_num = int(d.strftime('%y')) - int(gr_num[8:]) + int(d.month > 8)
    if course_num > 3 or course_num < 0:
        str = "Нет такой группы в расписании :("
        return str
    sheet = schedule_list[course_num - 1]
    num_cols = sheet.max_column
    group_cell = None
    if ndw == 7:
        str = "Выходной"
        return str
    for i in range(6, num_cols + 1, 5):
        if (sheet.cell(row=2, column=i).value == gr_num):
            group_cell = [2, i]
            break
    if group_cell == None:
        str = "Нет такой группы в расписании :("
        return str
    sub_cell = [4 + (ndw - 1) * 12, i]
    for j in range(2):
        for i in range(6):
            val = sheet.cell(sub_cell[0] + i * 2 + j, sub_cell[1]).value
            val_inf = [sheet.cell(sub_cell[0] + i * 2 + j, sub_cell[1] + 1).value,
                       sheet.cell(sub_cell[0] + i * 2 + j, sub_cell[1] + 2).value,
                       sheet.cell(sub_cell[0] + i * 2 + j, sub_cell[1] + 3).value]
            if val != None and str(val).lower() != str(val).upper():
                s[j] += str(i + 1) + ') '
                for t in range(len(val.split('\n'))):
                    s[j] += str(val).split('\n')[t]
                    print(val_inf)
                    for y in range(3):
                        if val_inf[y] == None:
                            s[j] += ", ---"
                        else:
                            try:
                                s[j] += ", " + str(val_inf[y]).split('\n')[t]
                            except IndexError:
                                pass
                    s[j] += '\n'
    for k in range(2):
        if k == 0:
            str += "Нечетная неделя, " + dw + ":\n"
        else:
            str += "Четная неделя, " + dw + ":\n"
        if s[k] == '':
            str += "Выходной"
        else:
            str += s[k]
    return str


# расписание на любую неделю
def schedule_for_any_week(d, gr_num):
    course_num = int(d.strftime('%y')) - int(gr_num[8:]) + int(d.month > 8)
    if course_num > 3 or course_num < 0:
        sh = "Нет такой группы в расписании :("
        return sh
    print(course_num)
    sheet = schedule_list[course_num - 1]
    gr_num = gr_num.upper()
    num_cols = sheet.max_column
    num_rows = sheet.max_row
    group_cell = None
    while (d.isoweekday() != 1):
        d += datetime.timedelta(days=-1)
    wn = int(d.isoweekday())
    sh = ''
    s = [''] * 7
    for i in range(6, num_cols + 1, 5):
        if (sheet.cell(row=2, column=i).value == gr_num):
            group_cell = [2, i]
            break
    if group_cell == None:
        sh = "Нет такой группы в расписании :("
        return sh
    sub_cell = [4 + (week_number(d) + 1) % 2, i]
    for j in range(6):
        for i in range(6):
            val = sheet.cell(j * 12 + sub_cell[0] + i * 2, sub_cell[1]).value
            val_inf = [sheet.cell(j * 12 + sub_cell[0] + i * 2, sub_cell[1] + 1).value,
                       sheet.cell(j * 12 + sub_cell[0] + i * 2, sub_cell[1] + 2).value,
                       sheet.cell(j * 12 + sub_cell[0] + i * 2, sub_cell[1] + 3).value]
            if val != None and str(val).lower() != str(val).upper():
                s[j] += str(i + 1) + ') '
                for t in range(len(val.split('\n'))):
                    s[j] += str(val).split('\n')[t]
                    for y in range(3):
                        if val_inf[y] == None:
                            s[j] += ", ---"
                        else:
                            try:
                                s[j] += ", " + str(val_inf[y]).split('\n')[t]
                            except IndexError:
                                pass
                    s[j] += '\n'
    for k in range(7):
        sh += "----------" + weekday[k + 1] + ', ' + (d + datetime.timedelta(days=k)).strftime("%d.%m") + '----------\n'
        if s[k] == '':
            sh += 'Выходной\n'
        else:
            sh += s[k]
    return sh


# возвращает список из преподавателей с фамилией name
def return_teachers(name):
    teachers = []
    for sheets in schedule_list:
        sheet = sheets
        num_cols = sheet.max_column
        num_rows = sheet.max_row
        for i in range(1, num_cols + 1):
            if sheet.cell(row=3, column=i).value == "ФИО преподавателя":
                for j in range(3, num_rows + 1):
                    if sheet.cell(row=j, column=i).value != None:
                        if '\n' in str(sheet.cell(row=j, column=i).value):
                            var = str(sheet.cell(row=j, column=i).value).split('\n')
                        else:
                            var = str(sheet.cell(row=j, column=i).value).split(', ')
                        for k in range(len(var)):
                            if name.lower() in var[k].lower() and var[k].lstrip().strip() not in teachers:
                                teachers.append(var[k].strip().lstrip())
    if teachers != []:
        return teachers
    else:
        return -1


# полное расписание для преподавателя
def get_shcedule_for_teacher(name):
    flag = 0
    rasp = [{}, {}]
    for i in range(2):
        for j in range(6):
            rasp[i][weekday[j + 1]] = ['- - - -'] * 6
    for sheets in schedule_list:
        sheet = sheets
        num_cols = sheet.max_column
        num_rows = sheet.max_row
        num_cols = sheet.max_column
        num_rows = sheet.max_row
        group = ''
        num_day, num_w, num_p = '', '', ''
        for i in range(1, num_cols + 1):
            if sheet.cell(row=3, column=i).value == "ФИО преподавателя":
                for j in range(3, num_rows + 1):
                    if sheet.cell(row=j, column=i).value != None:
                        if '\n' in str(sheet.cell(row=j, column=i).value):
                            var = str(sheet.cell(row=j, column=i).value).split('\n')
                        else:
                            var = str(sheet.cell(row=j, column=i).value).split(', ')
                        for k in range(len(var)):
                            if name.lower() == var[k].lower().strip().lstrip():
                                flag = 1
                                num_day = weekday[(j - 4) // 12 + 1]
                                num_w = (j) % 2
                                num_p = ((j - 4) % 12) // 2
                                group = str(sheet.cell(row=2, column=i - 2).value)
                                if rasp[num_w][num_day][num_p] != '- - - -':
                                    rasp[num_w][num_day][num_p] = group + ', ' + rasp[num_w][num_day][num_p]
                                else:
                                    try:
                                        rasp[num_w][num_day][num_p] = group + ' - ' + str(
                                            sheet.cell(row=j, column=i - 2).value.split('\n')[k])
                                        if str(sheet.cell(row=j, column=i - 1).value).split('\n')[k] != None:
                                            rasp[num_w][num_day][num_p] += ', ' + str(sheet.cell(row=j,
                                                                                                 column=i - 1).value).split(
                                                '\n')[k]
                                        if str(sheet.cell(row=j, column=i + 1).value).split('\n')[k] != None:
                                            rasp[num_w][num_day][num_p] += ', ' + str(sheet.cell(row=j,
                                                                                                 column=i + 1).value).split(
                                                '\n')[k]
                                    except IndexError:
                                        rasp[num_w][num_day][num_p] = group + ' - ' + str(
                                            sheet.cell(row=j, column=i - 2).value)
                                        if str(sheet.cell(row=j, column=i - 1).value) is not None:
                                            rasp[num_w][num_day][num_p] += ', ' + str(
                                                sheet.cell(row=j, column=i - 1).value)
                                        if str(sheet.cell(row=j, column=i + 1).value) is not None:
                                            rasp[num_w][num_day][num_p] += ', ' + str(
                                                sheet.cell(row=j, column=i + 1).value)
    if flag == 0:
        return -1
    return rasp


# отправка сообщения о состоянии погоды на текущий момент
def weather_now():
    weather_response = requests.get(
        'http://api.openweathermap.org/data/2.5/weather?q=moscow&appid=64bedaabfd9c3d473c77fd20bacc2458&units=metric&lang=ru')
    info = weather_response.json()
    print(info)
    icon_id = info["weather"][0]["icon"]  # иконка погоды
    req_ph = f'http://openweathermap.org/img/wn/' + icon_id + '@2x.png'
    ph = requests.get(req_ph, stream=True).raw
    send_photo(event.user_id, ph, "Погода в Москве")
    w = str(info["weather"][0]["description"].capitalize()) + ", температура: " + str(
        floor(info["main"]["temp_min"])) + " - " + str(ceil(info["main"]["temp_max"])) + "°С"
    w += "\nДавление: " + str(round(info["main"]["pressure"] / 1.333)) + " мм рт.ст., влажность: " + str(
        info["main"]["humidity"]) + "%"
    w += "\nВетер: " + wind_type[float(info['wind']['speed'])] + ", " + str(info['wind']['speed']) + " м/с, " + \
         wind_direction[float(info['wind']['deg'])]
    send_message(event.user_id, w)


# отправить сообщение - погода сегодня
def weather_today():
    w_res = requests.get(
        'http://api.openweathermap.org/data/2.5/forecast?lat=55.7522&lon=37.6156&appid=64bedaabfd9c3d473c77fd20bacc2458&units=metric&lang=ru')
    info = w_res.json()
    icons = []
    weather_info = [0] * 4
    br_inf, all_inf = '', ''
    for i in range(0, 9, 2):
        if str(info['list'][i]["dt_txt"][:10]) == str(datetime.date.today() + datetime.timedelta(days=1)) and int(
                info['list'][i]["dt_txt"][11:13]) > 5:
            break
        weather_info[day_time[(int(info['list'][i]["dt_txt"][11:13]))]] = info['list'][i]
    for j in range(len(weather_info)):
        i = weather_info[j]
        if i != 0:
            icon_id = i["weather"][0]["icon"]  # иконка погоды
            req_ph = f'http://openweathermap.org/img/wn/' + icon_id + '@2x.png'
            icons.append(requests.get(req_ph, stream=True))
            br_inf += '/' + str(round((float(i['main']['temp_min']) + float(i['main']['temp_min'])) / 2)) + '°С/'
            all_inf += day_time_name[j].upper() + '\n'
            all_inf += '//' + str(i["weather"][0]["description"].capitalize()) + ", температура: " + str(
                floor(i["main"]["temp_min"])) + " - " + str(ceil(i["main"]["temp_max"])) + "°С"
            all_inf += "\n//Давление: " + str(round(i["main"]["pressure"] / 1.333)) + " мм рт.ст., влажность: " + str(
                i["main"]["humidity"]) + "%"
            all_inf += "\n//Ветер: " + wind_type[float(i['wind']['speed'])] + ", " + str(i['wind']['speed']) + " м/с, " + \
                       wind_direction[float(i['wind']['deg'])] + '\n'
    img = Image.new('RGBA', (100 * len(icons), 100), color="grey")
    for i in range(len(icons)):  # создаем объединенную картинку погоды
        with open(f'file{i + 1}.jpeg', 'wb') as f:
            f.write(icons[i].content)
        paste_img = Image.open(f'file{i + 1}.jpeg')
        img.paste(paste_img, (i * 100, 0), paste_img)
    img.save("today.png")
    send_photo(event.user_id, "today.png", "Погода в Москве сегодня")
    send_message(event.user_id, br_inf)
    send_message(event.user_id, all_inf)


# погода завтра
def weather_tomorrow():
    w_res = requests.get(
        'http://api.openweathermap.org/data/2.5/forecast?lat=55.7522&lon=37.6156&appid=64bedaabfd9c3d473c77fd20bacc2458&units=metric&lang=ru')
    info = w_res.json()
    icons = []
    weather_info = [0] * 4
    br_inf, all_inf = '', ''
    for i in range(0, 17, 2):
        if str(info['list'][i]["dt_txt"][:10]) == str(datetime.date.today() + datetime.timedelta(days=2)) and int(
                info['list'][i]["dt_txt"][11:13]) > 5:
            break
        weather_info[day_time[(int(info['list'][i]["dt_txt"][11:13]))]] = info['list'][i]
    for j in range(len(weather_info)):
        i = weather_info[j]
        if i != 0:
            icon_id = i["weather"][0]["icon"]  # иконка погоды
            req_ph = f'http://openweathermap.org/img/wn/' + icon_id + '@2x.png'
            icons.append(requests.get(req_ph, stream=True))
            br_inf += '/' + str(round((float(i['main']['temp_min']) + float(i['main']['temp_min'])) / 2)) + '°С/'
            all_inf += day_time_name[j].upper() + '\n'
            all_inf += '//' + str(i["weather"][0]["description"].capitalize()) + ", температура: " + str(
                floor(i["main"]["temp_min"])) + " - " + str(ceil(i["main"]["temp_max"])) + "°С"
            all_inf += "\n//Давление: " + str(round(i["main"]["pressure"] / 1.333)) + " мм рт.ст., влажность: " + str(
                i["main"]["humidity"]) + "%"
            all_inf += "\n//Ветер: " + wind_type[float(i['wind']['speed'])] + ", " + str(i['wind']['speed']) + " м/с, " + \
                       wind_direction[float(i['wind']['deg'])] + '\n'
    img = Image.new('RGBA', (100 * len(icons), 100), color="grey")
    for i in range(len(icons)):  # создаем объединенную картинку погоды
        with open(f'file{i + 1}.jpeg', 'wb') as f:
            f.write(icons[i].content)
        paste_img = Image.open(f'file{i + 1}.jpeg')
        img.paste(paste_img, (i * 100, 0), paste_img)
    img.save("today.png")
    send_photo(event.user_id, "today.png", "Погода в Москве завтра")
    send_message(event.user_id, br_inf)
    send_message(event.user_id, all_inf)


# прогноз погоды на 5 дней


def weather_for_five_days():
    w_res = requests.get(
        'http://api.openweathermap.org/data/2.5/forecast?lat=55.7522&lon=37.6156&appid=64bedaabfd9c3d473c77fd20bacc2458&units=metric&lang=ru')
    info = w_res.json()
    days = {}
    br_inf, all_inf = '', ''
    for i in range(5):
        days[(datetime.date.today() + datetime.timedelta(days=i)).strftime('%Y-%m-%d')] = [0, 0]
    for i in range(len(info['list'])):
        if info['list'][i]["dt_txt"][:10] == (datetime.date.today() + datetime.timedelta(days=5)).strftime('%Y-%m-%d'):
            break
        if int(info['list'][i]["dt_txt"][11:13]) >= 6 and int(info['list'][i]["dt_txt"][11:13]) < 17:
            days[info['list'][i]["dt_txt"][:10]][0] = info['list'][i]
        if int(info['list'][i]["dt_txt"][11:13]) >= 21 or int(info['list'][i]["dt_txt"][11:13]) < 4:
            days[info['list'][i]["dt_txt"][:10]][1] = info['list'][i]
    morning, night = '', ''
    icons = []
    print(days)
    for i in days:
        if days[i][0] == 0:
            morning += '/---/'
            if days[i][1] != 0:
                icon_id = days[i][1]["weather"][0]["icon"]  # иконка погоды
                req_ph = f'http://openweathermap.org/img/wn/' + icon_id + '@2x.png'
                icons.append(requests.get(req_ph, stream=True))
        else:
            morning += '/' + str(ceil(days[i][0]["main"]["temp"])) + '°С/'
            icon_id = days[i][0]["weather"][0]["icon"]  # иконка погоды
            req_ph = f'http://openweathermap.org/img/wn/' + icon_id + '@2x.png'
            icons.append(requests.get(req_ph, stream=True))
        if days[i][1] == 0:
            night += '/---/'
        else:
            night += '/' + str(ceil(days[i][1]["main"]["temp"])) + '°С/'
    img = Image.new('RGBA', (100 * len(icons), 100), color="grey")
    for i in range(len(icons)):  # создаем объединенную картинку погоды
        with open(f'file-5days{i + 1}.jpeg', 'wb') as f:
            f.write(icons[i].content)
        paste_img = Image.open(f'file-5days{i + 1}.jpeg')
        img.paste(paste_img, (i * 100, 0), paste_img)
    img.save("5days.png")
    morning = morning + "ДЕНЬ"
    night = night + "НОЧЬ"
    send_photo(event.user_id, "5days.png", "Погода в Москве с " + (datetime.date.today()).strftime('%d.%m') + ' по ' + (
            datetime.date.today() + datetime.timedelta(days=4)).strftime('%d.%m'))
    send_message(event.user_id, morning + '\n' + night)


# статистика по ковиду в россии
def corona_rus():
    page = requests.get('https://coronavirusstat.ru/country/russia/')  # адрес старницы со статистикой
    soup = BeautifulSoup(page.text, "html.parser")
    # текстовое сообщение со статистикой
    information = "По состоянию на " + soup.find('strong').text + '\n'
    numbers = soup.find_all('div', class_="col col-6 col-md-3 pt-4")
    for i in numbers:
        number = i.find('b')
        information += number.find_next().text.lower() + ": " + number.text + " (" + i.contents[1].text.replace('(',
                                                                                                                'за ') + '\n'
    corona_graph()
    send_photo(event.user_id, 'covid.png', information)


# график
def corona_graph():
    page = requests.get('https://coronavirusstat.ru/country/russia/')  # адрес старницы со статистикой
    soup1 = BeautifulSoup(page.text, "html.parser")
    stat_info = soup1.find('table').find('tbody')
    dates = []
    arr1, arr2, arr3 = [], [], []
    for i in range(10):
        dates.append(stat_info.find_all('th')[i].text)
        arr1.append(int(stat_info.find_all('tr')[i].find_all('td')[0].text.split()[0]))  # активных
        arr2.append(arr1[i] + int(stat_info.find_all('tr')[i].find_all('td')[1].text.split()[0]))  # вылечено
        arr3.append(arr2[i] + int(stat_info.find_all('tr')[i].find_all('td')[2].text.split()[0]))  # умерло
    dates.reverse()
    arr1.reverse()
    arr2.reverse()
    arr3.reverse()
    matplotlib.use("TkAgg")
    fig, ax = plt.subplots()
    plt.plot(dates, arr1, "black")
    plt.plot(dates, arr2, "red")
    plt.ylim([0, 20000000])
    plt.title("Россия - детальная статистика - коронавирус")
    plt.plot(dates, arr3, label="умерло", color='red')
    plt.fill_between(dates, arr3, color='red')
    plt.plot(dates, arr2, label="вылечено", color='green')
    plt.fill_between(dates, arr2, color='green')
    plt.plot(dates, arr1, label="активных", color='yellow')
    plt.fill_between(dates, arr1, color='yellow')
    plt.xticks(rotation=25, fontsize=7)
    plt.yticks(fontsize=7)
    plt.gca().yaxis.set_major_formatter(mticker.FormatStrFormatter('%d'))
    plt.legend()
    plt.grid(True)
    fig.savefig('covid.png')


# статистика по региону
def corona_reg(rname):
    reg_name = check_name(rname)
    if reg_name == -1:
        send_message(event.user_id,
                     f"Регион с именем '{rname}' не был найден. Попробуйте ввести название региона иным способом.")
    else:
        page = requests.get(reg_dict[reg_name])  # адрес старницы со статистикой
        soup_reg = BeautifulSoup(page.text, "html.parser")
        information = "По состоянию на " + soup_reg.find('strong').text + '\nРегион: ' + reg_name + '\n'
        numbers = soup_reg.find_all('div', class_="col col-6 col-md-3 pt-4")
        for i in numbers:
            number = i.find('b')
            information += number.find_next().text.lower() + ": " + number.text + " (" + i.contents[1].text.replace('(',
                                                                                                                    'за ') + '\n'
        send_message(event.user_id, information)


# name - название региона вводимое пользователем, ф-ия изменяет название таким образом, чтобы оно соответствовало названию из словаря с регионами
def check_name(name):
    regions = []
    d = ['автономная обл.', 'обл.', 'край', 'республика', 'автономный округ', ')']
    for i in range(len(reg_list)):
        regions.append(reg_list[i].lower())
        for j in d:
            regions[i] = regions[i].replace(j, '')
        regions[i] = regions[i].replace(' — ', ' (').replace('  - ', ' (').lstrip().rstrip().split(' (')
    for i in range(len(regions)):
        for j in regions[i]:
            if j in name.lower():
                return (reg_list[i])
    return -1


# номер недели по дате (при отсутствии праметра возваращает номер текущей недели)
def week_number(d=datetime.date.today()):
    s = int(d.strftime("%V")) - int(datetime.date(2022, 2, 9).strftime("%V")) + 1
    return s


for event in longpoll.listen():
    if event.type == VkEventType.MESSAGE_NEW and event.to_me:
        print('New from {} with id {}, text = {}'.format(vk.users.get(user_ids=event.user_id)[0]['first_name'],
                                                         event.user_id, event.text))  # просто в консоль
        message_text = str(event.text).lower()

        if message_text == 'начать':
            send_message(event.user_id, strings.introduction)

            flag_user_in_user_list = 0  # ставим юзеру состояние старт или добавляем в список юзеров
            for user in users:
                if event.user_id == user.id:
                    user.mode = 'start'
                    flag_user_in_user_list = 1
                    break
            if flag_user_in_user_list == 0:
                users.append(User(event.user_id, "start", 0))
        for user in users:
            if user.id == event.user_id:
                if message_text.split(' ')[0] == 'группа':
                    user.group = (' '.join(message_text.split(' ')[1:])).upper()
                    send_message(user.id, "Название группы сохранено.")
                elif re.match(group_pattern, message_text) is not None:
                    user.group = message_text.upper()
                    users_dumped = dict(json_users)  # список пользователей
                    users_dumped = users_dumped.get("users")
                    if(users_dumped is None):
                        users_dumped = []
                    user_data = {"id": event.user_id, "mode": "start", "group": message_text.upper()}
                    for unique_user in users_dumped:
                        if unique_user.get("id")==user.id:
                            unique_user["group"]=message_text.upper()
                        else:
                            users_dumped.append(user_data)
                    data = {"users": users_dumped}
                    with open("users.json", "w") as f:
                        json.dump(data, f)
                    send_message(user.id, "Название группы сохранено.")
                elif user.group == 0:
                    send_message(event.user_id, "Чтобы продолжить работу, введите, пожалуйста, номер группы.")
                    user.mode = "get_group"
                elif message_text == 'бот':
                    user.mode = 'sсhedule'
                    keyboard = VkKeyboard(one_time=True)
                    keyboard.add_button("на сегодня", color=VkKeyboardColor.POSITIVE)
                    keyboard.add_button("на завтра", color=VkKeyboardColor.NEGATIVE)
                    keyboard.add_line()
                    keyboard.add_button("на эту неделю", color=VkKeyboardColor.PRIMARY)
                    keyboard.add_button("на следующую неделю", color=VkKeyboardColor.PRIMARY)
                    keyboard.add_line()
                    keyboard.add_button("какая неделя?", color=VkKeyboardColor.SECONDARY)
                    keyboard.add_button("какая группа?", color=VkKeyboardColor.SECONDARY)
                    send_message(event.user_id, "Показать расписание ...", keyboard)
                elif message_text.split(' ')[0] == 'бот':
                    if len(message_text.split(' ')) == 2:
                        if re.match(group_pattern, message_text.split(' ')[1]) is not None:
                            user.mode = 'schedule2'
                            user.search_group = message_text.split(' ')[1]
                            keyboard = VkKeyboard(one_time=False)
                            keyboard.add_button("на сегодня", color=VkKeyboardColor.POSITIVE)
                            keyboard.add_button("на завтра", color=VkKeyboardColor.NEGATIVE)
                            keyboard.add_line()
                            keyboard.add_button("на эту неделю", color=VkKeyboardColor.PRIMARY)
                            keyboard.add_button("на следующую неделю", color=VkKeyboardColor.PRIMARY)
                            keyboard.add_line()
                            keyboard.add_button("какая неделя?", color=VkKeyboardColor.SECONDARY)
                            keyboard.add_button("какая группа?", color=VkKeyboardColor.SECONDARY)
                            send_message(event.user_id, "Показать расписание ...", keyboard)
                        elif message_text.split(' ')[1] in weekday_list:
                            send_message(user.id, schedule_week_day(message_text.split(' ')[1], user.group))
                    elif len(message_text.split(' ')) == 3:
                        if message_text.split(' ')[1] in weekday_list and re.match(
                                group_pattern, message_text.split(' ')[2]) is not None:
                            send_message(user.id,
                                         schedule_week_day(message_text.split(' ')[1], message_text.split(' ')[2]))
                        elif message_text.split(' ')[2] in weekday_list and re.match(
                                group_pattern, message_text.split(' ')[1]) is not None:
                            send_message(user.id, "Возможно вы имелли в виду бот '" + message_text.split(' ')[2] + ' ' +
                                         message_text.split(' ')[1] + "'?")
                elif message_text == "какая группа?":
                    send_message(user.id, str(user.group))
                elif message_text == "погода":
                    user.mode = 'weather'
                    keyboard = VkKeyboard(one_time=True)
                    keyboard.add_button("сейчас", color=VkKeyboardColor.PRIMARY)
                    keyboard.add_button("сегодня", color=VkKeyboardColor.POSITIVE)
                    keyboard.add_button("завтра", color=VkKeyboardColor.POSITIVE)
                    keyboard.add_line()
                    keyboard.add_button("на 5 дней", color=VkKeyboardColor.POSITIVE)
                    send_message(event.user_id, "Выберите интересующий период", keyboard)
                elif message_text == "сейчас" and user.mode == 'weather':
                    weather_now()
                elif message_text == "сегодня" and user.mode == 'weather':
                    weather_today()
                elif message_text == "завтра" and user.mode == 'weather':
                    weather_tomorrow()
                elif message_text == "на 5 дней" and user.mode == 'weather':
                    weather_for_five_days()
                elif message_text == "коронавирус":
                    user.mode = 'corona'
                    try:
                        corona_rus()
                    except AttributeError:
                        send_message(user.id, "Не удалось получить информацию, попробуйте заново.")
                elif message_text.split(' ')[0] == "коронавирус":
                    user.mode = 'corona'
                    corona_reg(' '.join(message_text.split(' ')[1:]))
                elif message_text == 'какая неделя?':
                    send_message(event.user_id, "Идет " + str(week_number()) + "-ая учебная неделя")
                elif message_text.split(' ')[0] == "найти" and message_text.split(' ')[0] != message_text:
                    t_names = return_teachers(' '.join(message_text.split(' ')[1:]))
                    if t_names == -1:
                        send_message(user.id, "Не был найден такой преподаватель")
                    else:
                        user.mode = 'search'
                        if len(t_names) > 1:
                            keyboard = VkKeyboard(one_time=False)
                            for i in range(len(t_names)):
                                keyboard.add_button(t_names[i], color=VkKeyboardColor.PRIMARY)
                                user.mode = 'get_teacher_name'
                            send_message(user.id, "Выберите преподавателя ...", keyboard)
                        else:
                            keyboard = VkKeyboard(one_time=False)
                            user.search_teacher = t_names[0]
                            keyboard.add_button("на сегодня", color=VkKeyboardColor.POSITIVE)
                            keyboard.add_button("на завтра", color=VkKeyboardColor.NEGATIVE)
                            keyboard.add_line()
                            keyboard.add_button("на эту неделю", color=VkKeyboardColor.PRIMARY)
                            keyboard.add_button("на следующую неделю", color=VkKeyboardColor.PRIMARY)
                            send_message(user.id, "Выберите ...", keyboard)
                elif user.mode == "get_teacher_name":
                    user.search_teacher = message_text
                    user.mode = 'search'
                    keyboard = VkKeyboard(one_time=False)
                    keyboard.add_button("на сегодня", color=VkKeyboardColor.POSITIVE)
                    keyboard.add_button("на завтра", color=VkKeyboardColor.NEGATIVE)
                    keyboard.add_line()
                    keyboard.add_button("на эту неделю", color=VkKeyboardColor.PRIMARY)
                    keyboard.add_button("на следующую неделю", color=VkKeyboardColor.PRIMARY)
                    send_message(user.id, "Выберите ...", keyboard)
                elif user.mode == "search":
                    s = ''
                    ras = get_shcedule_for_teacher(user.search_teacher)
                    if ras == -1:
                        send_message(user.id, 'Ошибка. Преподаватель не найден.')
                    else:
                        if message_text == 'на сегодня':
                            d = datetime.date.today()
                            s += "Расписание на " + d.strftime(
                                '%d.%m') + " для преподавателя " + user.search_teacher.capitalize()[0:len(
                                user.search_teacher) - 4] + user.search_teacher.capitalize()[
                                                            len(user.search_teacher) - 4:].upper() + '\n'
                            for k in range(6):
                                s += str(k + 1) + ") " + ras[(week_number(d) + 1) % 2][weekday[int(d.isoweekday())]][
                                    k] + '\n'
                            send_message(user.id, s)
                        elif message_text == "на завтра":
                            d = datetime.date.today() + datetime.timedelta(days=1)
                            s += "Расписание на " + d.strftime(
                                '%d.%m') + " для преподавателя " + user.search_teacher.capitalize()[0:len(
                                user.search_teacher) - 4] + user.search_teacher.capitalize()[
                                                            len(user.search_teacher) - 4:].upper() + '\n'
                            for k in range(6):
                                s += str(k + 1) + ") " + ras[(week_number(d) + 1) % 2][weekday[int(d.isoweekday())]][
                                    k] + '\n'
                            send_message(user.id, s)
                        elif message_text == "на эту неделю":
                            s += "Расписание на эту неделю для преподавателя " + user.search_teacher.capitalize()[0:len(
                                user.search_teacher) - 4] + user.search_teacher.capitalize()[
                                                            len(user.search_teacher) - 4:].upper() + '\n'
                            d = datetime.date.today()
                            for i in range(1, 7):
                                s += '---' + weekday[i] + '---\n'
                                for j in range(6):
                                    s += str(j + 1) + ") " + ras[(week_number(d) + 1) % 2][weekday[i]][j] + '\n'
                            send_message(user.id, s)
                        elif message_text == "на следующую неделю":
                            s += "Расписание на следующую неделю для преподавателя " + user.search_teacher.capitalize()[
                                                                                       0:len(
                                                                                           user.search_teacher) - 4] + user.search_teacher.capitalize()[
                                                                                                                       len(user.search_teacher) - 4:].upper() + '\n'
                            d = datetime.date.today() + datetime.timedelta(weeks=1)
                            for i in range(1, 7):
                                s += '----------' + weekday[i] + '----------\n'
                                for j in range(6):
                                    s += str(j + 1) + ") " + ras[(week_number(d) + 1) % 2][weekday[i]][j] + '\n'
                            send_message(user.id, s)
                        else:
                            send_message(user.id, strings.unknown)
                elif user.mode == "sсhedule" or user.mode == 'schedule2':
                    if user.mode == "sсhedule":
                        gruppa = str(user.group)
                    else:
                        gruppa = user.search_group
                    if message_text == "на сегодня":
                        d = datetime.date.today()
                        send_message(user.id, schedule_for_any_day(gruppa, d))
                    elif message_text == "на завтра":
                        d = datetime.date.today() + datetime.timedelta(days=1)
                        send_message(user.id, schedule_for_any_day(gruppa, d))
                    elif message_text == "на эту неделю":
                        d = datetime.date.today()
                        send_message(user.id, schedule_for_any_week(d, gruppa))
                    elif message_text == "на следующую неделю":
                        d = datetime.date.today() + datetime.timedelta(weeks=1)
                        send_message(user.id, schedule_for_any_week(d, gruppa))
                    else:
                        send_message(user.id, strings.unknown)
                elif message_text == 'помощь':
                    send_message(event.user_id, strings.help)

                else:
                    if message_text != "начать":
                        send_message(user.id, strings.unknown)
